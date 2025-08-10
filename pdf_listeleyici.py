import os
import platform
import subprocess
import re
from tkinter import filedialog, Tk, messagebox
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime
import tkinter as tk
import urllib.parse

APP_ROOT = None

def temizle(metin):
    if metin is None:
        return ""
    if not isinstance(metin, str):
        metin = str(metin)
    return re.sub(r'[\x00-\x1F\x7F-\x9F]', '', metin)

def excel_ac(dosya_yolu):
    try:
        if platform.system() == "Windows":
            os.startfile(dosya_yolu)
        elif platform.system() == "Darwin":  # MacOS
            subprocess.call(["open", dosya_yolu])
        else:  # Linux
            subprocess.call(["xdg-open", dosya_yolu])
    except Exception as e:
        print(f"Excel dosyası açılamadı: {e}")

def pdf_bilgileri_al(dosya_yolu):
    try:
        reader = PdfReader(dosya_yolu)
        sayfa_sayisi = len(reader.pages) if reader.pages is not None else 0
        meta = reader.metadata or {}

        tarih_raw = meta.get('/CreationDate', None)
        tarih = "Bilinmiyor"
        if tarih_raw:
            tarih_str = tarih_raw[2:] if isinstance(tarih_raw, str) and tarih_raw.startswith("D:") else str(tarih_raw)
            tarih_match = re.match(r"(\d{4})(\d{2})(\d{2})(\d{2})(\d{2})(\d{2})", tarih_str)
            if tarih_match:
                yil, ay, gun, saat, dakika, saniye = tarih_match.groups()
                try:
                    tarih_dt = datetime(int(yil), int(ay), int(gun), int(saat), int(dakika), int(saniye))
                    tarih = tarih_dt.strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    tarih = tarih_str

        try:
            mtime = os.path.getmtime(dosya_yolu)
            son_degistirme = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            son_degistirme = ""

        yazar = meta.get('/Author', None) or meta.get('Author', 'Bilinmiyor')
        baslik = meta.get('/Title', None) or meta.get('Title', 'Bilinmiyor')
        uretici = meta.get('/Producer', None) or meta.get('Producer', 'Bilinmiyor')
        konu = meta.get('/Subject', None) or meta.get('Subject', 'Bilinmiyor')

        ilk_sayfa_metni = ""
        if sayfa_sayisi > 0:
            try:
                metin = reader.pages[0].extract_text()
                ilk_sayfa_metni = metin or ""
            except Exception:
                ilk_sayfa_metni = ""

        ilk_300_karakter = temizle(ilk_sayfa_metni[:300]) if ilk_sayfa_metni else "Metin yok"

        try:
            boyut = os.path.getsize(dosya_yolu)
        except Exception:
            boyut = None

        return {
            "dosya_adi": temizle(os.path.basename(dosya_yolu)),
            "dosya_yolu": os.path.abspath(dosya_yolu),
            "sayfa": sayfa_sayisi,
            "boyut_bytes": boyut,
            "tarih": temizle(tarih),
            "son_degistirme": son_degistirme,
            "yazar": temizle(yazar),
            "baslik": temizle(baslik),
            "uretici": temizle(uretici),
            "konu": temizle(konu),
            "ilk_satirlar": ilk_300_karakter
        }
    except Exception as e:
        print(f"Hata {dosya_yolu} → {e}")
        return None

def collect_pdf_paths(klasor_listesi, recursive=True):
    paths = []
    for klasor in klasor_listesi:
        if recursive:
            for root, dirs, files in os.walk(klasor):
                for f in files:
                    if f.lower().endswith(".pdf"):
                        paths.append(os.path.join(root, f))
        else:
            try:
                for f in os.listdir(klasor):
                    if f.lower().endswith(".pdf"):
                        paths.append(os.path.join(klasor, f))
            except Exception as e:
                print(f"Klasör okunamadı: {klasor} → {e}")
    return sorted(paths)

def excel_yaz_coklu_klasor(veriler, kaydet_yolu):
    wb = Workbook()
    ws = wb.active

    basliklar = ["Klasör", "Dosya Adı", "PDF Açmak için Link", "Sayfa Sayısı", "Oluşturulma Tarihi", "Yazar", "Başlık", "Üretici", "Konu", "İlk Satırlar"]
    ws.append(basliklar)

    baslik_font = Font(bold=True, color="FFFFFF")
    baslik_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    kenarlik = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    wrap_alignment = Alignment(wrap_text=True, vertical='top')

    for col_num, baslik in enumerate(basliklar, 1):
        hucre = ws.cell(row=1, column=col_num)
        hucre.font = baslik_font
        hucre.fill = baslik_fill
        hucre.border = kenarlik
        hucre.alignment = Alignment(horizontal='center', vertical='center')

    for satir_index, veri in enumerate(veriler, start=2):
        ws.cell(row=satir_index, column=1, value=veri.get("klasor", "")).border = kenarlik
        ws.cell(row=satir_index, column=1).alignment = wrap_alignment

        ws.cell(row=satir_index, column=2, value=veri["dosya_adi"]).border = kenarlik
        ws.cell(row=satir_index, column=2).alignment = wrap_alignment

        link_hucre = ws.cell(row=satir_index, column=3, value="PDF Aç")
        try:
            dosya_yolu = veri["dosya_yolu"].replace("\\", "/")
            uri = f"file:///{dosya_yolu}"
            link_hucre.hyperlink = uri
            link_hucre.style = "Hyperlink"
            link_hucre.font = Font(color="0000FF", underline="single")
        except Exception:
             link_hucre.value = veri["dosya_yolu"]

        link_hucre.border = kenarlik
        link_hucre.alignment = Alignment(horizontal='center', vertical='center')


        ws.cell(row=satir_index, column=4, value=veri["sayfa"]).border = kenarlik
        ws.cell(row=satir_index, column=4).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=satir_index, column=5, value=veri["tarih"]).border = kenarlik
        ws.cell(row=satir_index, column=5).alignment = wrap_alignment

        ws.cell(row=satir_index, column=6, value=veri["yazar"]).border = kenarlik
        ws.cell(row=satir_index, column=6).alignment = wrap_alignment

        ws.cell(row=satir_index, column=7, value=veri["baslik"]).border = kenarlik
        ws.cell(row=satir_index, column=7).alignment = wrap_alignment

        ws.cell(row=satir_index, column=8, value=veri["uretici"]).border = kenarlik
        ws.cell(row=satir_index, column=8).alignment = wrap_alignment

        ws.cell(row=satir_index, column=9, value=veri["konu"]).border = kenarlik
        ws.cell(row=satir_index, column=9).alignment = wrap_alignment

        ws.cell(row=satir_index, column=10, value=veri["ilk_satirlar"]).border = kenarlik
        ws.cell(row=satir_index, column=10).alignment = wrap_alignment

        if satir_index % 2 == 0:
            fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
            for col in range(1, 11):
                ws.cell(row=satir_index, column=col).fill = fill

        ws.row_dimensions[satir_index].height = 40

    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 35
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 35
    ws.column_dimensions['J'].width = 70

    wb.save(kaydet_yolu)



def klasor_sec_ve_rapor_olustur():
    global APP_ROOT
    if APP_ROOT is None:
        # Eğer APP_ROOT yoksa oluştur
        APP_ROOT = Tk()
        APP_ROOT.withdraw()

    secilen_klasorler = []

    top = tk.Toplevel(APP_ROOT)
    top.title("Çoklu Klasör Seçimi")
    top.geometry("720x380")
    top.transient(APP_ROOT)

    label = tk.Label(top, text="PDF klasörlerini seçin (her seferinde bir tane). İşlem bittiğinde 'İşlem Bitti' butonuna basın.", wraplength=700)
    label.pack(pady=8)

    recursive_var = tk.BooleanVar(value=True)
    chk_recursive = tk.Checkbutton(top, text="Alt klasörlerde de ara (recursive)", variable=recursive_var)
    chk_recursive.pack()

    btn_frame = tk.Frame(top)
    btn_frame.pack(pady=6)

    def klasor_sec():
        klasor = filedialog.askdirectory(title="PDF klasörü seçin", parent=top)
        if klasor:
            secilen_klasorler.append(klasor)
            listbox_klasorler.insert(tk.END, klasor)

    btn_klasor = tk.Button(btn_frame, text="Klasör Ekle", command=klasor_sec)
    btn_klasor.pack(side=tk.LEFT, padx=4)

    btn_temizle = tk.Button(btn_frame, text="Temizle", command=lambda: (secilen_klasorler.clear(), listbox_klasorler.delete(0, tk.END)))
    btn_temizle.pack(side=tk.LEFT, padx=4)

    listbox_klasorler = tk.Listbox(top, width=110, height=12)
    listbox_klasorler.pack(pady=8)

    def islem_bitir():
        top.destroy()

    btn_bitir = tk.Button(top, text="İşlem Bitti", command=islem_bitir)
    btn_bitir.pack(pady=6)

    top.grab_set()
    top.wait_window()

    if not secilen_klasorler:
        messagebox.showwarning("Uyarı", "Hiç klasör seçilmedi!", parent=APP_ROOT)
        return

    all_pdf_paths = collect_pdf_paths(secilen_klasorler, recursive=recursive_var.get())
    if not all_pdf_paths:
        messagebox.showinfo("Bilgi", "Seçilen klasör(ler)de PDF bulunamadı.", parent=APP_ROOT)
        return

    toplam_pdf = len(all_pdf_paths)
    toplam_boyut = 0
    for p in all_pdf_paths:
        try:
            toplam_boyut += os.path.getsize(p) or 0
        except:
            pass
    toplam_mb = round(toplam_boyut / (1024 * 1024), 2)

    devam = messagebox.askyesno("Onay", f"{len(secilen_klasorler)} klasör seçildi.\nToplam PDF: {toplam_pdf}\nToplam boyut: {toplam_mb} MB\n\nİşleme başlamak istiyor musunuz?", parent=APP_ROOT)
    if not devam:
        return

    veriler = []
    error_list = []

    progress = tk.Toplevel(APP_ROOT)
    progress.title("İşlem sürüyor...")
    lbl = tk.Label(progress, text="0 / {}".format(toplam_pdf))
    lbl.pack(padx=20, pady=20)
    progress.transient(APP_ROOT)
    progress.grab_set()
    progress.update_idletasks()

    for idx, pdf_path in enumerate(all_pdf_paths, start=1):
        info = pdf_bilgileri_al(pdf_path)
        if info:
            parent_folder = None
            for s in secilen_klasorler:
                if os.path.commonpath([s, pdf_path]) == s:
                    parent_folder = s
                    break
            info["klasor"] = parent_folder or os.path.dirname(pdf_path)
            veriler.append(info)
        else:
            error_list.append(pdf_path)

        lbl.config(text=f"{idx} / {toplam_pdf}")
        lbl.update_idletasks()

    progress.destroy()

    kaydet_yolu = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Dosyası", "*.xlsx")], title="Raporu Kaydet", parent=APP_ROOT)
    if not kaydet_yolu:
        messagebox.showinfo("Bilgi", "Kaydetme iptal edildi.", parent=APP_ROOT)
        return

    try:
        excel_yaz_coklu_klasor(veriler, kaydet_yolu)
    except PermissionError:
        return

    if error_list:
        log_path = os.path.splitext(kaydet_yolu)[0] + "_errors.log"
        with open(log_path, "w", encoding="utf-8") as f:
            for e in error_list:
                f.write(f"{e}\n")
        messagebox.showwarning("Uyarı", f"{len(error_list)} dosya işlenemedi. Hata listesi: {log_path}", parent=APP_ROOT)
    else:
        messagebox.showinfo("Başarılı", f"Excel raporu oluşturuldu:\n{kaydet_yolu}", parent=APP_ROOT)

    excel_ac(kaydet_yolu)

def gui_baslat():
    global APP_ROOT
    root = tk.Tk()
    APP_ROOT = root
    root.title("PDF Listeleyici")
    root.geometry("720x380")

    label = tk.Label(root, text="PDF klasörlerini seçin (her seferinde bir tane). İşlem bittiğinde 'İşlem Bitti' butonuna basın.", wraplength=700)
    label.pack(pady=8)

    btn_klasor = tk.Button(root, text="Klasör Ekle ve Rapor Oluştur", command=klasor_sec_ve_rapor_olustur)
    btn_klasor.pack(pady=8)

    root.mainloop()

if __name__ == "__main__":
    gui_baslat()
